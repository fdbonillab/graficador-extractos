from itertools import cycle
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt

#cargamos el xlsx en memoria
wb = load_workbook('datos.xlsx', data_only=True)
#cargamos la hoja
sheet1 = wb.get_sheet_by_name('Hoja1')

#creamos la matriz en numpy
matriz = np.zeros((sheet1.max_row, sheet1.max_column))

#leemos la hoja fila por fila y vamos guardando los valores 'x' e 'y'
for i in range(0,sheet1.max_row):
    for j in range(0,sheet1.max_column):
        matriz[i,j]=sheet1.cell(row=i+1, column=j+1).value

#creamos el grafico
colors = cycle(["aqua", "black", "blue", "fuchsia", "gray", "green", "lime", "maroon", "navy", "olive", "purple", "red", "silver", "teal", "yellow"])
plt.xlabel('tiempo')
plt.ylabel('intensidad')
for y in range(1, sheet1.max_column):
    plt.plot(matriz[:,0],matriz[:,y], label="Data " + str(y),       color=next(colors))
plt.legend(loc='upper left', fontsize='small')
plt.grid(True)
plt.xlim(0,70)
plt.ylim(0,70)
plt.title('Grafica tiempo/intensidad')
plt.show()
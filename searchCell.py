from openpyxl import Workbook
import openpyxl
import re
import os
import matplotlib.pyplot as plt
from itertools import cycle
from dateutil import parser 
import datetime
from statistics import mean 
#from sklearn.model_selection import train_test_split
#from sklearn.linear_model import LinearRegression
import datetime as dt
import numpy as np
import pandas as pd

#wb = load_workbook('datos.xlsx', data_only=True)

# fecha descripcion sucursal valor saldo
class EstadoDia:
  def __init__(self, fecha, descripcion, sucursal,valor, saldo, anio):
    self.fecha = fecha
    self.anio = anio
    self.descripcion = descripcion
    self.sucursal = sucursal
    self.valor = valor
    self.saldo = saldo
  def __init__(self, fecha, descripcion, sucursal,valor, saldo):
    self.fecha = fecha
    self.descripcion = descripcion
    self.sucursal = sucursal
    self.valor = valor
    self.saldo = saldo
  def __init__(self, fecha, descripcion, sucursal="",valor=0, saldo=0, anio=1996):
    self.fecha = fecha
    self.anio = anio
    self.descripcion = descripcion
    self.sucursal = sucursal
    self.valor = valor
    self.saldo = saldo

  def __repr__(self):
      return " fecha :"+self.fecha+" descripcion :"+self.descripcion+" valor :"+self.valor+" saldo "+self.saldo
  def __str__(self):
      return " fecha :"+self.fecha+" descripcion :"+self.descripcion+" valor :"+self.valor+" saldo "+self.saldo
  def mostrarFechaValor(self):
      print(" fecha : "+str(self.fecha)+" valor "+self.valor)

#p1 = EstadoDia("John", 36)
def cambiarFormatoAnio( strEntrada ):
    #format = '%b %d %Y %I:%M%p'
    #datetime_str = datetime.datetime.strptime(date_time, format) 
    #DT = parser.parse("Jun 23 2022 07:31PM") 
    #print(' formato entrada '+strEntrada)
    DT = parser.parse(strEntrada) 
    #datetime.datetime.strptime("21/12/2008", "%d/%m/%Y").strftime("%Y-%m-%d")
    formatoCambiado = datetime.datetime.strptime(strEntrada, "%d/%m/%Y").strftime("%Y-%m-%d")
    #print(' formato cambiado '+formatoCambiado) 
    #print(DT)
    return str(formatoCambiado)
def extraerAnio():
    nombreCelda = "DESDE"
    celdaHasta = "HASTA"
    anioHasta = 0
    mes = 0
    for row in ws.iter_rows(min_row=6,  
                            max_row=8,  
                            min_col=0,  
                            max_col=2):
        for cell in row:
            if cell.value == nombreCelda:
                ##print(ws.cell(row=(cell.row)+1, column=1).value) #change column number for any cell value you want
                mes = (ws.cell(row=(cell.row)+1, column=1).value)[5:7]
                ##print(' mes para extraccion '+mes)
                anio = (ws.cell(row=(cell.row)+1, column=1).value)[0:4]
               
            if cell.value == celdaHasta:
                anioHasta = (ws.cell(row=(cell.row)+1, column=2).value)[0:4]
                print(' anio hasta '+str(anioHasta))
    print(' mes para extraccion '+mes)        
    if( mes == '12' ):
        anio = anioHasta
    if anio is None:
        print("no se encontro nada") 
    return anio
def isDiciembre(fecha):
    #print(' fecha para mes '+fecha)
    mes = fecha.split("/")[1]
    #print(' mes '+mes)
    if( mes == 12 ):
        return True
    return False

def extraerFechas():
    arrFechas = []
    nombreCelda = "FECHA"
    fechaEncontrada = False
    for row in ws.iter_rows(min_row=6,  
                            max_row=1000,  
                            min_col=0,  
                            max_col=6):
        for cell in row:
            if cell.value == nombreCelda:
                fechaEncontrada = True
                ##print(ws.cell(row=(cell.row)+1, column=1).value) #change column number for any cell value you want
            if cell.value :    
                fecha = (ws.cell(row=(cell.row), column=1).value)
            if fechaEncontrada and fecha != None and cell.value :
                #print(fecha)
                indice = 0
                #print(' cell row '+str(cell.row))
                if has_numbers(fecha): 
                    #print(fecha)
                    descripcion = (ws.cell(row=(cell.row), column=2).value)
                    sucursal = ""
                    if  ws.cell(row=(cell.row), column=3).value != None:
                        sucursal = ws.cell(row=(cell.row), column=3).value
                    valor = ws.cell(row=(cell.row), column=5).value
                    if has_numbers(valor):
                        #print('valor '+valor)
                        valor = valor.split(".")[0]
                        #print('valor '+valor)
                        valor = valor.replace(",","")
                        #print('valor '+valor)
                        if len(valor) > 1 :
                            valor = str(round(float(valor)))
                            saldo = ws.cell(row=(cell.row), column=6).value
                            isDiciembre(fecha)
                            fecha = fecha+"/"+anio
                            fecha = cambiarFormatoAnio(fecha)
                            fecha = fecha[0:10]
                            estadoDia = EstadoDia(fecha=fecha,descripcion=descripcion,sucursal=sucursal,valor=valor,saldo=saldo)
                            agregarEstado(arrEstados=arrFechas,estado=estadoDia)
                            fecha = (ws.cell(row=(cell.row), column=1).value)
                            fecha = fecha+"/"+anio
                            indice = indice +1
                else :
                    continue
                #print(fecha+' salida ')
                
    return arrFechas
def agregarEstado( arrEstados,estado ):
    yaExiste = False
    for elEstado in arrEstados:
        if elEstado.fecha == estado.fecha and elEstado.descripcion == estado.descripcion and elEstado.valor == estado.valor:
            yaExiste = True
    if yaExiste == False:
        arrEstados.append(estado)
    if estado.fecha == '2024-03-25':
        print(' revisando error comparacion ')
        print(elEstado)
        print(yaExiste)
    return arrEstados
def greet():
    print('Hello World!')

def has_numbers(inputString):
    return bool(re.search(r'\d', inputString))
def pintarPrediccion():
    #xColumn= X_train[:,1] 
    xColumn = np.take(X_train, 0, axis=1)
    # Prediction on training set
    print(' len x '+str(len(X_train))+' len y '+str(len(y_train))+" len xtodos "+str(len(xTodos)))
    print(X_train)
    print(y_train)
    plt.scatter(xColumn, y_train, color = 'lightcoral')
    plt.plot(xColumn, y_pred_train, color = 'firebrick')
    plt.title('Salary vs Experience (Training Set)')
    plt.xlabel('Years of Experience')
    plt.ylabel('Salary')
    plt.legend(['X_train/Pred(y_test)', 'X_train/y_train'], title = 'Sal/Exp', loc='best', facecolor='white')
    plt.box(False)
    plt.show()
def armarDataFrame(matrixSaldos):
    columns = ["fecha", "valor"]
    rows = ["D", "E", "F"]
    data = matrixSaldos
    df = pd.DataFrame(data=data, index=range(0,len(matrixSaldos)), columns=columns)
    print('dataFrame ')
    print(df)
    return df
def armarDataFrameDesangre(listaDesangre):
    columns = ["fecha", "valor"]
    rows = ["D", "E", "F"]
    #parser.parse(strEntrada) 
    fechas = [o.fecha for o in listaDesangre]
    attrs = [o.valor for o in listaDesangre]
    descs = [o.descripcion for o in listaDesangre]
    print(' len attrs '+str(len(attrs)))
    matrix = [[fechas],[attrs]]
    #print(matrix)
    matrix = np.reshape(matrix, (-1,2))
    #matrix.shape = (2,55)
    #print(matrix)
    #df = pd.DataFrame(data=matrix, index=range(0,len(attrs)), columns=columns)
    #df = pd.DataFrame({'fecha': fechas, 'valor':attrs,'descripcion':descs}) 
    df = pd.DataFrame({'fecha': fechas, 'valor':attrs}) 
    print('dataFrame ')
    print(df)
    df['month'] = pd.to_datetime(df['fecha']).dt.month
    df['year'] = pd.to_datetime(df['fecha']).dt.year
    df = pd.concat([df['year'],df['month'],df['valor']], axis=1)
    print(df)
    promedioMensual = df.groupby(['year','month'],as_index=False).sum()
    print('suma mensual desangre')
    print(promedioMensual)
    return df 
def promedioMensual(df):
    df['month'] = pd.to_datetime(df['fecha']).dt.month
    df['year'] = pd.to_datetime(df['fecha']).dt.year
    promedioMensual = df.groupby(['year','month'],as_index=False).mean()
    print('promedio mensual')
    print(promedioMensual)
    mostrarPromedioX4Meses(promedioMensual)
   
    df['fec2'] = df['fecha'].astype(str).str[:7]
    #dfFecha = df['year']+''+df['month']
    plt.plot(df['fec2'], df['valor'],linewidth=3)
    plt.show()
def mostrarPromedioX4Meses( promedioMensual):
    #df1 = df[['a', 'b']]
    df1 = promedioMensual[['valor']]
    proCuatri = df1.groupby(df1.index / 4).mean()
    print(' len cuatri 1 a '+str(len(proCuatri)))
    columns = ['valor']
    longDF = len(proCuatri)-proCuatri.isna().sum()
    print(longDF)
    print(proCuatri.isna().sum())
    proCuatri = pd.DataFrame(data=proCuatri, index=range(0,8), columns=columns)
    #proCuatri.index = range(0,len(proCuatri))
    proCuatri2 = proCuatri.groupby(proCuatri.index / 4).mean()
    #proCuatri2.index = range(0,len(proCuatri2))
    #proCuatri['index'] = 
    print(' len cuatri 1 '+str(len(proCuatri)))
    print(' len cuatri 2 '+str(len(proCuatri2)))
    print(proCuatri)
    print(proCuatri2)
    primerPromedio = proCuatri[0:4].mean()
    segundoPromedio = proCuatri[4:8].mean()
    arr2Promedios = []
    arr2Promedios.append(primerPromedio)
    arr2Promedios.append(segundoPromedio)
    print(proCuatri[0:4])
    print(proCuatri[4:8])
    print(' primer promedio '+str(primerPromedio))
    print(' segundo promedio '+str(segundoPromedio))

    plt.plot(range(0,len(proCuatri)), proCuatri['valor'],linewidth=3)
    plt.show()
    plt.plot(range(0,len(proCuatri2)), proCuatri2['valor'],linewidth=3, color='red')
    plt.show()
    plt.plot([0,32], arr2Promedios,linewidth=3, color='red')
    plt.show()
    '''acum4 = 0
    cont = 0
    for cuatrimestre in promedioMensual:
        print(cuatrimestre)
        acum4 = cuatrimestre[3]+acum4
        cont = cont +1
        if cont == 4:
            print( 'promedio cuatrimestre '+acum4/4)
            cont = 0
            acum4 = 0'''
            

##greet()
files = [f for f in os.listdir() if os.path.isfile(f)]
print(files)
#file = "datos.xlsx"#"enter_path_to_file_here"
reporteTotal = []
reporteGeneral = []
contador = 0
ejeX = []
ejeY = []
current_time = datetime.datetime.now() 
cambiarFormatoAnio('29/03/2023')
for file in files:
    # se cambia a veces nombre de archivo para hacer pruebas pequeñas
    if file.endswith('.xlsx') or file.endswith('datos.xlsmr'):
        print(' leyendo archivo '+file)
        ##### break para pruebas cortas
        #break
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        anio = extraerAnio()
        print(anio+' extraido')
        ##### break para pruebas cortas
        #break
        reporteArchivo = extraerFechas()
        print( ' fechas extraidas ')
        for dato in reporteArchivo:
            algo = 1
            #dato.mostrarFechaValor()
        print (' lontitud arreglo archivo '+str(len(reporteArchivo)))
        reporteTotal.append(reporteArchivo)
        print (' lontitud arreglo reporte total '+str(len(reporteTotal)))
        contador= contador +1
        #break para solo procesar un archivo
        #break
print(' duracion ejecucion hasta aki '+str(datetime.datetime.now()-current_time))
for rep in reporteTotal:
    for subRep in rep:
        #reporteGeneral.append(subRep)
        agregarEstado(arrEstados=reporteGeneral,estado=subRep)
        #ejeX.append(subRep.fecha)
        #ejeY.append(int(subRep.valor))
print(' duracion ejecucion hasta aki '+str(datetime.datetime.now()-current_time))
reporteGeneral.sort(key=lambda x: x.fecha, reverse=False)
menoresA10mil = 0
mayoresA10mil = 0
arrMayores10mil = []
# Splitting dataset into test/train

paraPromedio = []
xTodos = []
yTodos = []
listaTransfSucursalVirtual = []
listaTransfDesangre = []
# se agregan los reportes de cada archivo a una sola lista
for rep in reporteGeneral:
        #ejeX.append(rep.fecha)
        xTodos.append(rep.fecha)
        valorNum = int(rep.valor)
        if abs(valorNum) < 10000:
            menoresA10mil = menoresA10mil+1
        else:
            mayoresA10mil = mayoresA10mil+1
            arrMayores10mil.append(rep)
        #ejeY.append(valorNum)
        yTodos.append(valorNum)
        paraPromedio.append(valorNum)
        if rep.descripcion == 'TRANSFERENCIA CTA SUC VIRTUAL':
            listaTransfSucursalVirtual.append(rep)
print(' duracion ejecucion hasta aki '+str(datetime.datetime.now()-current_time))
for transf in listaTransfSucursalVirtual:
    valorNum = abs(int(transf.valor))
    transf.valor = valorNum
    if ( valorNum > 500000 and valorNum < 800000) or ( valorNum > 15000 and valorNum < 30000):
        listaTransfDesangre.append(transf)

# contar las veces los valores mensual y promedio de las trasferencias de sucursal virtual
print(' transferencias suc virtual '+str(len(listaTransfSucursalVirtual)))
print(' transferencias desangre '+str(len(listaTransfDesangre)))
armarDataFrameDesangre(listaTransfDesangre)
#print(listaTransfSucursalVirtual)
#print(listaTransfDesangre)
print(' transferencias desangre sum '+str(sum(c.valor for c in listaTransfDesangre) ) )
print(' promedio todos los valores '+str(mean(paraPromedio)))

yDependiente = []
matrixParaDataFrame = []
for i in range (0, len(xTodos)):
    #print(xTodos[i])
    dateFecha =  datetime.datetime.strptime(xTodos[i], '%Y-%m-%d')
    #print('datefecha')
    #print(dateFecha)
    #print(' timestamp ')
    #print(dateFecha.timestamp())
    xTodos[i] = int(dateFecha.timestamp())
    union =[ xTodos[i], yTodos[i]]
    yDependiente.append(union)
    matrixParaDataFrame.append([dateFecha,yTodos[i]])
df = armarDataFrame(matrixParaDataFrame)
promedioMensual(df)

#xTodos= xTodos.map(dt.datetime.toordinal)
# Splitting dataset into test/train
X_train, X_test, y_train, y_test = train_test_split(yDependiente, yTodos, test_size = 0.2, random_state = 0)
# Regressor model
regressor = LinearRegression()
regressor.fit(X_train, y_train)
# Prediction result
y_pred_test = regressor.predict(X_test)     # predicted value of y_test
y_pred_train = regressor.predict(X_train)   # predicted value of y_train
#aki se pinta la prediccion
#pintarPrediccion()

for rep in arrMayores10mil:
    ejeX.append(rep.fecha)
    ejeY.append(int(rep.valor))
print (' conteo reporte general '+str(len(reporteGeneral)))
print(' menores a 10k '+str(menoresA10mil))
print(' mayores a 10k '+str(mayoresA10mil))
#creamos el grafico
colors = cycle(["aqua", "black", "blue", "fuchsia", "gray", "green", "lime", "maroon", "navy", "olive", "purple", "red", "silver", "teal", "yellow"])
plt.xlabel('tiempo')
plt.ylabel('valor')
#subX = ejeX[0:5]
#subY = ejeY[0:5]
subX = ejeX
subY = ejeY
#print(subX)
#print(subY)
print(' min suby '+str(min(subY)))
print(' max suby '+str(max(subY)))
current_time2 = datetime.datetime.now() 
print(' duracion ejecucion hasta aki '+str(current_time2-current_time))
for rep in arrMayores10mil:
    estado = rep
    #plt.plot(ejeX, ejeY, label="Data " + str(estado.fecha),       color=next(colors))
    #plt.plot(ejeX, ejeY, 'rs')
    plt.plot(subX, subY, 'rs')
    plt.ylim(min(subY), max(subY))
    #plt.xlim(min(subX), max(subX))
    #plt.plot([1, 2, 3,4], [1, 4, -19,23], 'rs')
plt.legend(loc='upper left', fontsize='small')
plt.grid(True)
#plt.xlim(0,70)
#plt.ylim(0,70)
plt.title('Grafica tiempo/valor')
plt.show()
